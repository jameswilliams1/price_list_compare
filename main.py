"""Program entry point."""
import pandas as pd
import glob
import os
from functools import reduce
import sys
import openpyxl as xl
from openpyxl.styles import Alignment


def compare(df_1, df_2):
    """Join two dataframes to give a comparison of all prices."""
    df = pd.merge(df_1, df_2, on="part", how="outer")
    return df.sort_values("part")


def get_frames_from_dir(path):
    """Return a list of (name, dataframe) tuples from a directory.

    Uses current directory if none specified.
    """
    all_files = glob.glob(os.path.join(path, "*.xlsx"))
    data_list = [
        (
            os.path.basename(filename).split(".")[0],
            pd.read_excel(filename, index_col="part"),
        )
        for filename in all_files
    ]
    return sorted(data_list, key=lambda tup: tup[0])


def merge_frame_list(frame_list):
    """Merge a list of (name: frame) tuples using compare.

    Returns final processed dataframe.
    """
    return reduce(compare, map(lambda x: x[1], frame_list))


def write_file(df, file_name):
    """Write a dataframe to an xlsx file."""
    writer = pd.ExcelWriter(f"{file_name}", engine="openpyxl")
    df.to_excel(writer, sheet_name="Sheet1", startrow=1)
    writer.save()


def cleanup(filename, frames):
    """Add company names from frames above data and tidy column names."""
    workbook = xl.load_workbook(filename)
    ws = workbook.active
    alignment = Alignment(horizontal="center", vertical="bottom")
    for i in range(2, 2 * len(frames) + 1, 2):
        ws.cell(row=1, column=i).value = frames[int(i / 2) - 1][0]
        ws.cell(row=1, column=i).alignment = alignment
        ws.merge_cells(
            start_row=1, start_column=i, end_row=1, end_column=i + 1
        )
    for i in range(2, 2 * len(frames) + 2):
        ws.cell(row=2, column=i).value = ws.cell(row=2, column=i).value.split(
            "_"
        )[0]
    all_rows = ws.iter_rows(
        min_row=1, min_col=1, max_col=1 + 2 * len(frames), max_row=5000
    )
    for row in all_rows:
        for col in row:
            col.alignment = alignment
    workbook.save(filename)


def main(path=""):
    """Write a joined xlsx file from a directory of xlsx files.

    If no directory specified, uses the current one.
    """
    filename = "price_list_compare.xlsx"
    frames = get_frames_from_dir(path=path)
    if not frames:
        print("No spreadsheets found in specified directory.")
        sys.exit(1)
    merged = merge_frame_list(frames)
    write_file(merged, filename)
    cleanup(filename, frames)


if __name__ == "__main__":
    main("")
